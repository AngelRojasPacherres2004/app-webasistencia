from __future__ import annotations

import json
from collections import defaultdict
from datetime import date, datetime, timedelta
from html import escape
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _finish_workbook(workbook: Workbook, sheet) -> bytes:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(
            max_length + 4, 42
        )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _format_time(value) -> str:
    if value in (None, "", False) or str(value).strip() in {"None", "--"}:
        return ""
    text = str(value).strip()
    if "T" in text:
        text = text.split("T", 1)[-1]
    if " " in text:
        text = text.split(" ", 1)[-1]
    return text[:5]


def _report_date(value) -> date:
    return datetime.fromisoformat(str(value)[:10]).date()


def _calculate_report_hours(row: dict) -> tuple[str, str]:
    def parse_minutes(value) -> int | None:
        text = _format_time(value)
        if not text:
            return None
        try:
            parsed = datetime.strptime(text, "%H:%M")
            return parsed.hour * 60 + parsed.minute
        except ValueError:
            return None

    entry_minutes = parse_minutes(row.get("hora_inicio"))
    exit_minutes = parse_minutes(row.get("hora_final"))
    if entry_minutes is None or exit_minutes is None:
        return "", ""

    exit_floor_minutes = (exit_minutes // 60) * 60
    total_minutes = exit_floor_minutes - entry_minutes
    has_break = bool(
        _format_time(row.get("inicio_receso"))
        and _format_time(row.get("final_receso"))
    )
    if has_break:
        total_minutes -= 60
    total_minutes = max(total_minutes, 0)
    hours, minutes = divmod(total_minutes, 60)
    calculation = (
        f"{exit_floor_minutes // 60:02d}:00 - "
        f"{_format_time(row.get('hora_inicio')) or '-'}"
    )
    if has_break:
        calculation += " - 1:00 receso"
    calculation += f" = {hours}:{minutes:02d}"
    return f"{hours}:{minutes:02d}", calculation


def attendance_excel(rows: list[dict], metadata: dict) -> bytes:
    """Detailed attendance workbook matching the original panel export."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen"

    title_fill = PatternFill("solid", fgColor="1E3A5F")
    title_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    header_font = Font(color="1E3A5F", bold=True)

    sheet["A1"] = "Reporte de Asistencias"
    sheet["A1"].fill = title_fill
    sheet["A1"].font = title_font
    sheet["A1"].alignment = Alignment(horizontal="center")

    start = _report_date(metadata.get("start"))
    end = _report_date(metadata.get("end"))
    metadata_rows = [
        ("Periodo", metadata.get("label") or "-"),
        ("Rango", f"{start:%d/%m/%Y} - {end:%d/%m/%Y}"),
        ("Tienda", metadata.get("store_label") or "Todas las tiendas"),
        ("Persona", metadata.get("worker_label") or "Todas"),
        ("Búsqueda", metadata.get("search") or "-"),
        ("Generado", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for row_index, (label, value) in enumerate(metadata_rows, 3):
        sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row_index, column=2, value=value)

    headers = [
        "#",
        "Trabajador",
        "DNI",
        "Sede",
        "Fecha",
        "Entrada",
        "Ini. receso",
        "Fin receso",
        "Salida",
        "Horas",
        "Cálculo horas",
        "Tardanza",
        "Justificado",
    ]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(row=10, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for index, row in enumerate(rows, 1):
        hours, calculation = _calculate_report_hours(row)
        sheet.append(
            [
                index,
                row.get("nombre_trabajador") or "-",
                row.get("dni") or "-",
                row.get("nombre_tienda") or "-",
                row.get("fecha") or "-",
                _format_time(row.get("hora_inicio")),
                _format_time(row.get("inicio_receso")),
                _format_time(row.get("final_receso")),
                _format_time(row.get("hora_final")),
                hours,
                calculation,
                "Sí" if row.get("late") else "No",
                "Sí" if row.get("justificado") else "No",
            ]
        )

    if rows:
        sheet.freeze_panes = "A11"
        sheet.auto_filter.ref = f"A10:M{sheet.max_row}"
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(
            max_length + 4, 40
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def marks_excel(rows: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Marcas"
    headers = [
        "#",
        "Trabajador",
        "DNI",
        "Tienda",
        "Fecha",
        "Hora",
        "Tipo",
        "Ubicación",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for index, row in enumerate(rows, 1):
        location = row.get("ubicacion")
        if isinstance(location, (dict, list)):
            location = json.dumps(location, ensure_ascii=False)
        sheet.append(
            [
                index,
                row.get("nombre_trabajador") or "-",
                row.get("id_trabajador") or "-",
                row.get("nombre_tienda") or "-",
                row.get("fecha_local") or "-",
                row.get("hora_local") or "-",
                row.get("tipo") or "-",
                location or "-",
            ]
        )
    return _finish_workbook(workbook, sheet)


def attendance_pdf(rows: list[dict], workers: list[dict], metadata: dict) -> bytes:
    """Weekly landscape matrix matching the original panel PDF."""
    workers = sorted(
        workers,
        key=lambda worker: str(worker.get("nombre_trabajador") or ""),
    )
    black = colors.HexColor("#0F172A")
    dark = colors.HexColor("#1E293B")
    muted = colors.HexColor("#64748B")
    light = colors.HexColor("#F1F5F9")
    header_background = colors.HexColor("#1E3A5F")
    row_alternate = colors.HexColor("#F8FAFC")
    accent = colors.HexColor("#2563EB")
    green_text = colors.HexColor("#166534")
    blue_text = colors.HexColor("#1D4ED8")
    red_background = colors.HexColor("#DC2626")
    late_background = colors.HexColor("#FEF3C7")
    missing_background = colors.HexColor("#F1F5F9")

    def style(
        name,
        size=8,
        color=None,
        bold=False,
        align=None,
        leading=10,
        background=None,
    ):
        options = {
            "fontName": "Helvetica-Bold" if bold else "Helvetica",
            "fontSize": size,
            "textColor": color or black,
            "leading": leading,
        }
        if align is not None:
            options["alignment"] = align
        if background is not None:
            options["backColor"] = background
        return ParagraphStyle(name, **options)

    title_style = style("attendance_title", size=16, bold=True, leading=20)
    subtitle_style = style(
        "attendance_subtitle", size=8, color=dark, leading=11
    )
    footer_style = style(
        "attendance_footer",
        size=6.5,
        color=muted,
        align=TA_RIGHT,
        leading=9,
    )
    meta_label_style = style("attendance_meta_label", size=7.5, bold=True)
    meta_value_style = style(
        "attendance_meta_value", size=7.5, color=dark
    )
    table_header_style = style(
        "attendance_table_header",
        size=7.5,
        color=colors.white,
        bold=True,
        align=TA_CENTER,
    )
    table_header_left_style = style(
        "attendance_table_header_left",
        size=7.5,
        color=colors.white,
        bold=True,
    )
    week_style = style(
        "attendance_week", size=8, color=colors.HexColor("#1E40AF"), bold=True
    )
    worker_style = style("attendance_worker", size=7, bold=True)
    number_style = style(
        "attendance_number", size=7, color=muted, align=TA_CENTER
    )
    on_time_style = style(
        "attendance_on_time",
        size=6.5,
        color=green_text,
        bold=True,
        align=TA_CENTER,
        background=colors.HexColor("#DCFCE7"),
    )
    late_style = style(
        "attendance_late",
        size=6.5,
        color=colors.HexColor("#92400E"),
        bold=True,
        align=TA_CENTER,
        background=late_background,
    )
    justified_style = style(
        "attendance_justified",
        size=6.5,
        color=blue_text,
        bold=True,
        align=TA_CENTER,
        background=colors.HexColor("#DBEAFE"),
    )
    absent_style = style(
        "attendance_absent",
        size=6.5,
        color=colors.white,
        bold=True,
        align=TA_CENTER,
        background=red_background,
    )
    missing_style = style(
        "attendance_missing",
        size=6.5,
        color=muted,
        bold=True,
        align=TA_CENTER,
        background=missing_background,
    )

    def attendance_cell(row: dict | None):
        if row is None:
            return Paragraph("FALTA", absent_style)
        parts = []
        if row.get("justificado") and row.get("fuera_horario"):
            parts.append("JUSTIFICADO")
        parts.extend(
            value
            for value in (
                _format_time(row.get("hora_inicio")),
                _format_time(row.get("inicio_receso")),
                _format_time(row.get("final_receso")),
                _format_time(row.get("hora_final")),
            )
            if value
        )
        if not parts:
            return Paragraph("—", missing_style)
        content = "<br/>".join(escape(value) for value in parts)
        if row.get("justificado") and row.get("fuera_horario"):
            return Paragraph(content, justified_style)
        return Paragraph(content, late_style if row.get("late") else on_time_style)

    start = _report_date(metadata.get("start"))
    end = _report_date(metadata.get("end"))
    attendance_index: dict[str, dict[date, dict]] = defaultdict(dict)
    for row in rows:
        try:
            row_date = date.fromisoformat(str(row.get("fecha", ""))[:10])
        except ValueError:
            continue
        attendance_index[str(row.get("dni") or "").strip()][row_date] = row

    weeks = []
    cursor = start - timedelta(days=start.weekday())
    while cursor <= end:
        days = [
            cursor + timedelta(days=offset)
            for offset in range(6)
            if start <= cursor + timedelta(days=offset) <= end
        ]
        if days:
            weeks.append(days)
        cursor += timedelta(days=7)

    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")
    late_count = sum(1 for row in rows if row.get("late"))
    on_time_count = len(rows) - late_count
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    story = [
        Paragraph("Reporte de Asistencias", title_style),
        Paragraph(
            (
                f"Periodo: <b>{escape(str(metadata.get('label') or '-'))}</b> · "
                f"Rango: {start:%d/%m/%Y} — {end:%d/%m/%Y} · "
                f"Generado: {generated_at}"
            ),
            subtitle_style,
        ),
        HRFlowable(
            width="100%",
            thickness=1.2,
            color=accent,
            spaceAfter=2 * mm,
        ),
    ]

    store_label = escape(str(metadata.get("store_label") or "Todas las tiendas"))
    search_label = escape(str(metadata.get("search") or "—"))
    meta_rows = [
        [
            Paragraph("<b>Tienda</b>", meta_label_style),
            Paragraph(store_label, meta_value_style),
            Paragraph("<b>Trabajadores</b>", meta_label_style),
            Paragraph(str(len(workers)), meta_value_style),
            Paragraph("<b>Registros</b>", meta_label_style),
            Paragraph(str(len(rows)), meta_value_style),
        ],
        [
            Paragraph("<b>Búsqueda</b>", meta_label_style),
            Paragraph(search_label, meta_value_style),
            Paragraph("<b>A tiempo</b>", meta_label_style),
            Paragraph(str(on_time_count), meta_value_style),
            Paragraph("<b>Tardanzas</b>", meta_label_style),
            Paragraph(str(late_count), meta_value_style),
        ],
    ]
    meta_table = Table(
        meta_rows,
        colWidths=[22 * mm, 68 * mm, 28 * mm, 22 * mm, 24 * mm, 22 * mm],
    )
    meta_table.setStyle(
        TableStyle(
            [
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, light]),
                ("BOX", (0, 0), (-1, -1), 0.5, muted),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, muted),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.extend([meta_table, Spacer(1, 3 * mm)])

    legend = Table(
        [
            [
                Paragraph("<b>Leyenda de colores:</b>", meta_label_style),
                Paragraph("Puntual", on_time_style),
                Paragraph("Tardanza", late_style),
                Paragraph("Justificado", justified_style),
                Paragraph("Falta", absent_style),
                Paragraph("Sin marca", missing_style),
            ]
        ],
        colWidths=[34 * mm, 32 * mm, 34 * mm, 34 * mm, 28 * mm, 30 * mm],
    )
    legend.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.4, muted),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 0), (-1, -1), light),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.extend([legend, Spacer(1, 4 * mm)])

    day_names = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
    for week_number, week_days in enumerate(weeks, 1):
        story.append(
            Paragraph(
                f"Semana {week_number} · "
                f"{week_days[0]:%d/%m} — {week_days[-1]:%d/%m}",
                week_style,
            )
        )
        day_width = (246 - 8 - 42) / len(week_days)
        table_rows = [
            [
                Paragraph("N°", table_header_style),
                Paragraph("Trabajador", table_header_left_style),
                *[
                    Paragraph(
                        f"{day_names[day.weekday()]}<br/>{day:%d/%m}",
                        table_header_style,
                    )
                    for day in week_days
                ],
            ]
        ]
        for index, worker in enumerate(workers, 1):
            worker_attendance = attendance_index.get(
                str(worker.get("dni") or "").strip(), {}
            )
            table_rows.append(
                [
                    Paragraph(str(index), number_style),
                    Paragraph(
                        escape(str(worker.get("nombre_trabajador") or "-")),
                        worker_style,
                    ),
                    *[
                        attendance_cell(worker_attendance.get(day))
                        for day in week_days
                    ],
                ]
            )
        week_table = Table(
            table_rows,
            colWidths=[8 * mm, 42 * mm]
            + [day_width * mm] * len(week_days),
            repeatRows=1,
        )
        week_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), header_background),
                    ("BOX", (0, 0), (-1, -1), 0.5, muted),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, muted),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, row_alternate],
                    ),
                ]
            )
        )
        story.extend([week_table, Spacer(1, 5 * mm)])

    story.extend(
        [
            HRFlowable(width="100%", thickness=0.4, color=muted),
            Paragraph(
                f"Generado el {generated_at} | {len(workers)} trabajadores | "
                f"{len(rows)} registros | {on_time_count} a tiempo | "
                f"{late_count} tardanzas",
                footer_style,
            ),
        ]
    )
    document.build(story)
    return buffer.getvalue()


def workers_pdf(workers: list[dict], store_label: str, query: str) -> bytes:
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Directorio de trabajadores", styles["Title"]),
        Paragraph(
            f"Tienda: {escape(store_label or 'Todas')} · "
            f"Búsqueda: {escape(query or '—')} · Registros: {len(workers)}",
            styles["Normal"],
        ),
        Spacer(1, 5 * mm),
    ]
    rows = [
        [
            "#",
            "Nombre",
            "DNI",
            "Cargo",
            "Tienda",
            "Correo",
            "Teléfono",
            "Estado",
        ]
    ]
    for index, worker in enumerate(workers, 1):
        rows.append(
            [
                index,
                worker.get("nombre_trabajador") or "-",
                worker.get("dni") or "-",
                worker.get("area") or "-",
                worker.get("nombre_sede") or "-",
                worker.get("correo") or "-",
                worker.get("telefono") or "-",
                "Activo" if worker.get("estado") else "Inactivo",
            ]
        )
    table = Table(rows, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.35,
                    colors.HexColor("#CBD5E1"),
                ),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F8FAFC")],
                ),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(table)
    document.build(story)
    return buffer.getvalue()
