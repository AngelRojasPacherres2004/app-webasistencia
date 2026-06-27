import json
import sys
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

import streamlit as st

sys.path.append(str(Path(__file__).parent.parent))
from config.db import get_pooled_connection


LIMA_TZ = ZoneInfo("America/Lima")


# ================================================================
#  HELPERS DB
# ================================================================

def _get_tiendas():
    with get_pooled_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id_tienda, nombre, direccion FROM public.tienda ORDER BY nombre")
            return cur.fetchall()


def _get_trabajadores(id_tienda=None):
    with get_pooled_connection() as conn:
        with conn.cursor() as cur:
            if id_tienda:
                cur.execute(
                    "SELECT dni, nombre FROM public.trabajador WHERE id_tienda::text = %s AND estado = true ORDER BY nombre",
                    (id_tienda,),
                )
            else:
                cur.execute("SELECT dni, nombre FROM public.trabajador WHERE estado = true ORDER BY nombre")
            return cur.fetchall()


def _get_asistencias(fecha_inicio, fecha_fin, id_tienda=None, id_trabajador=None):
    inicio_local = datetime.combine(fecha_inicio, datetime.min.time(), tzinfo=LIMA_TZ)
    fin_local = datetime.combine(fecha_fin + timedelta(days=1), datetime.min.time(), tzinfo=LIMA_TZ)
    with get_pooled_connection() as conn:
        with conn.cursor() as cur:
            query = """
                SELECT
                    am.id,
                    am.id_tienda,
                    t.nombre AS nombre_tienda,
                    t.direccion AS direccion_tienda,
                    am.id_trabajador,
                    tr.nombre AS nombre_trabajador,
                    am.hora_marca,
                    (am.hora_marca AT TIME ZONE 'America/Lima') AS hora_marca_local,
                    DATE(am.hora_marca AT TIME ZONE 'America/Lima') AS fecha_local,
                    am.ubicacion,
                    am.tipo
                FROM public.asistencia_multiple am
                LEFT JOIN public.tienda t ON t.id_tienda::text = am.id_tienda::text
                LEFT JOIN public.trabajador tr ON tr.dni = am.id_trabajador
                WHERE am.hora_marca >= %s
                  AND am.hora_marca < %s
            """
            params = [inicio_local, fin_local]

            if id_tienda:
                query += " AND am.id_tienda::text = %s"
                params.append(id_tienda)
            if id_trabajador:
                query += " AND am.id_trabajador = %s"
                params.append(id_trabajador)

            query += " ORDER BY am.hora_marca DESC"
            cur.execute(query, params)
            return cur.fetchall()


def _export_excel(rows):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    headers = [
        "#",
        "Nombre del trabajador",
        "DNI",
        "Tienda",
        "Fecha",
        "Hora",
        "Tipo",
        "Ubicacion",
    ]
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, 1):
        hora_marca = row.get("hora_marca_local") or row.get("hora_marca")
        if isinstance(hora_marca, datetime):
            fecha_str = hora_marca.strftime("%d/%m/%Y")
            hora_str = hora_marca.strftime("%H:%M:%S")
        elif isinstance(hora_marca, str):
            fecha_str = hora_marca[:10]
            hora_str = hora_marca[11:19] if len(hora_marca) >= 19 else hora_marca
        else:
            fecha_str = "-"
            hora_str = "-"

        ubicacion = row.get("ubicacion")
        if isinstance(ubicacion, dict):
            ubicacion = json.dumps(ubicacion, ensure_ascii=True)

        ws.append(
            [
                i,
                row.get("nombre_trabajador") or "-",
                row.get("id_trabajador") or "-",
                row.get("nombre_tienda") or "-",
                fecha_str,
                hora_str,
                row.get("tipo") or "-",
                ubicacion or "-",
            ]
        )

    if ws.max_row > 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ================================================================
#  VISTA PRINCIPAL
# ================================================================

def render_asistencias_multiples(api=None):
    st.markdown(
        """
        <div style="margin-bottom:24px;">
            <h2 class="page-title">Asistencias</h2>
            <p class="page-subtitle">Historial de marcas de asistencia por trabajador y tienda</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Reutiliza los datos ya cargados por el panel para evitar otra consulta.
    try:
        tiendas = (
            [
                {
                    "id_tienda": store.get("id_tienda", ""),
                    "nombre": store.get("nombre_tienda", ""),
                    "direccion": store.get("direccion", ""),
                }
                for store in api.get_tiendas()
            ]
            if api is not None
            else _get_tiendas()
        )
    except Exception as e:
        st.error(f"Error al conectar con la base de datos: {e}")
        return

    tienda_map = {t["nombre"]: str(t["id_tienda"]) for t in tiendas}

    # Filtros
    with st.container(border=True):
        col1, col2, col3, col4, col5 = st.columns([2, 2, 1.5, 1.5, 1])

        tienda_options = ["Todas"] + list(tienda_map.keys())
        selected_tienda = col1.selectbox("Tienda", options=tienda_options, key="am_tienda")

        id_tienda_filter = tienda_map.get(selected_tienda) if selected_tienda != "Todas" else None

        try:
            if api is not None:
                trabajadores = [
                    {
                        "dni": worker.get("dni", ""),
                        "nombre": worker.get("nombre_trabajador", ""),
                    }
                    for worker in api.get_trabajadores()
                    if worker.get("estado", True)
                    and (
                        not id_tienda_filter
                        or str(worker.get("id_sede", "")) == str(id_tienda_filter)
                    )
                ]
            else:
                trabajadores = _get_trabajadores(id_tienda_filter)
        except Exception:
            trabajadores = []

        trabajador_map = {t["nombre"]: t["dni"] for t in trabajadores}
        trabajador_options = ["Todos"] + list(trabajador_map.keys())
        selected_trabajador = col2.selectbox("Persona", options=trabajador_options, key="am_trabajador")

        fecha_inicio = col3.date_input("Desde", value=date.today() - timedelta(days=7), key="am_desde")
        fecha_fin = col4.date_input("Hasta", value=date.today(), key="am_hasta")
        buscar = col5.button("Buscar", use_container_width=True, key="am_buscar")

    if not buscar and "am_resultados" not in st.session_state:
        st.info("Selecciona los filtros y presiona Buscar.")
        return

    # Ejecutar busqueda
    if buscar:
        dni_filter = trabajador_map.get(selected_trabajador) if selected_trabajador != "Todos" else None
        try:
            rows = _get_asistencias(fecha_inicio, fecha_fin, id_tienda_filter, dni_filter)
            st.session_state["am_resultados"] = rows
            st.session_state["am_pagina"] = 0
        except Exception as e:
            st.error(f"Error al cargar asistencias: {e}")
            return
    else:
        rows = st.session_state["am_resultados"]

    if not rows:
        st.info("No se encontraron registros con los filtros seleccionados.")
        return

    # Exportar Excel
    excel_buf = _export_excel(rows)
    if excel_buf:
        nombre_archivo = f"asistencias_resumen_{fecha_inicio}_{fecha_fin}.xlsx"
        st.download_button(
            label="Exportar Excel",
            data=excel_buf,
            file_name=nombre_archivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )

    # Paginacion
    PAGE_SIZE = 15
    total = len(rows)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    pagina = st.session_state.get("am_pagina", 0)

    page_rows = rows[pagina * PAGE_SIZE : (pagina + 1) * PAGE_SIZE]

    st.caption(f"Mostrando {pagina * PAGE_SIZE + 1}-{min((pagina + 1) * PAGE_SIZE, total)} de {total} registros")

    # Tabla
    headers = ["#", "TRABAJADOR", "DNI", "TIENDA", "FECHA", "HORA", "TIPO", "UBICACION"]
    col_widths = [0.4, 2.5, 1, 2, 1.2, 1, 1.2, 1.5]
    head_cols = st.columns(col_widths)
    for col, h in zip(head_cols, headers):
        col.markdown(f"<span class='table-header'>{h}</span>", unsafe_allow_html=True)

    st.markdown("<hr style='margin:0.6rem 0; border-color:#e2e8f0;'>", unsafe_allow_html=True)

    for i, row in enumerate(page_rows, pagina * PAGE_SIZE + 1):
        if i > pagina * PAGE_SIZE + 1:
            st.markdown("<hr style='margin:0.3rem 0; border-color:#f1f5f9; opacity:0.6;'>", unsafe_allow_html=True)

        cols = st.columns(col_widths)

        hora_marca = _row_datetime(row)
        fecha_str = hora_marca.strftime("%d/%m/%Y") if hora_marca else "-"
        hora_str = hora_marca.strftime("%H:%M:%S") if hora_marca else "-"
        nombre_trab = row.get("nombre_trabajador") or "-"
        nombre_tnd = row.get("nombre_tienda") or "-"
        tipo = row.get("tipo", "-")
        direccion = row.get("direccion_tienda") or "-"

        if tipo == "MULTIPLE":
            badge = '<span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:99px;font-size:11px;font-weight:600;">MULTIPLE</span>'
        else:
            badge = '<span style="background:#d1fae5;color:#065f46;padding:2px 8px;border-radius:99px;font-size:11px;font-weight:600;">NORMAL</span>'

        ubi_raw = row.get("ubicacion")
        maps_url = None
        if ubi_raw:
            try:
                ubi = ubi_raw if isinstance(ubi_raw, dict) else json.loads(ubi_raw)
                lat = float(ubi.get("latitud", 0))
                lng = float(ubi.get("longitud", 0))
                if lat != 0 or lng != 0:
                    maps_url = f"https://maps.google.com/?q={lat},{lng}"
            except Exception:
                pass

        cols[0].caption(str(i))
        cols[1].markdown(f"<div class='row-main'>{nombre_trab}</div>", unsafe_allow_html=True)
        cols[2].caption(row.get("id_trabajador", "-"))
        cols[3].caption(nombre_tnd)
        cols[4].caption(fecha_str)
        cols[5].caption(hora_str)
        cols[6].markdown(badge, unsafe_allow_html=True)
        if maps_url:
            cols[7].markdown(f"[Marca]({maps_url})", unsafe_allow_html=True)
            cols[7].caption(f" {direccion}")
        else:
            cols[7].caption("Sin GPS")
            cols[7].caption(f" {direccion}")

    # Controles paginacion
    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    p1, p2, p3 = st.columns([1, 2, 1])
    if p1.button("Anterior", disabled=pagina == 0, use_container_width=True):
        st.session_state["am_pagina"] = pagina - 1
        st.rerun()
    p2.markdown(
        f"<div style='text-align:center;padding-top:0.4rem;font-size:0.85rem;'>Pagina {pagina + 1} de {total_pages}</div>",
        unsafe_allow_html=True,
    )
    if p3.button("Siguiente", disabled=pagina >= total_pages - 1, use_container_width=True):
        st.session_state["am_pagina"] = pagina + 1
        st.rerun()
