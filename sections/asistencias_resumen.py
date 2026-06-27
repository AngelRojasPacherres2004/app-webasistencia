from calendar import monthrange
from datetime import date, datetime, timedelta
import re

import streamlit as st
import streamlit.components.v1 as components


# ================================================================
#  CONSTANTES
# ================================================================

MONTH_NAMES = (
    "Enero","Febrero","Marzo","Abril","Mayo","Junio",
    "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre",
)

DAY_NAME_MAP = {
    0:"lunes", 1:"martes", 2:"miercoles", 3:"jueves",
    4:"viernes", 5:"sabado", 6:"domingo",
}


# ================================================================
#  HELPERS GENERALES
# ================================================================

def _parse_date(value):
    try:
        return date.fromisoformat(str(value or "")[:10])
    except ValueError:
        return None


def _week_start(current_date):
    return current_date - timedelta(days=current_date.weekday())


def _week_label(start_date):
    end_date = start_date + timedelta(days=6)
    return f"{start_date.strftime('%d/%m/%Y')} -> {end_date.strftime('%d/%m/%Y')}"


def _month_label(d):
    return f"{MONTH_NAMES[d.month - 1]} {d.year}"


def _safe_text(value):
    return str(value or "").replace("\r", " ").replace("\n", " ").strip()


def _sanitize_filename(text):
    safe = re.sub(r"[^a-zA-Z0-9._-]+", "_", _safe_text(text).lower()).strip("_")
    return safe or "reporte"


def _parse_time_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if "T" in text: text = text.split("T", 1)[-1]
        if " " in text: text = text.split(" ", 1)[-1]
        return text[:5]
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    return str(value)[:5]


def _time_value_to_minutes(value):
    text = _parse_time_value(value)
    if not text:
        return None
    try:
        parsed = datetime.strptime(text[:5], "%H:%M").time()
        return parsed.hour * 60 + parsed.minute
    except ValueError:
        return None


def _is_on_time_or_early(actual_value, scheduled_value):
    actual_minutes = _time_value_to_minutes(actual_value)
    scheduled_minutes = _time_value_to_minutes(scheduled_value)
    if actual_minutes is None or scheduled_minutes is None:
        return None
    return actual_minutes <= scheduled_minutes


def _as_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "t", "yes", "y", "si", "sí"}


def _scheduled_entry_for_row(row, schedule_map):
    parsed = _parse_date(row.get("fecha"))
    if not parsed:
        return None
    dni = str(row.get("dni", "")).strip()
    day_name = DAY_NAME_MAP.get(parsed.weekday(), "")
    return schedule_map.get(dni, {}).get(day_name)


def _is_justificable_row(row, schedule_map):
    if not row:
        return False
    if _as_bool(row.get("justificado")):
        return False
    return _scheduled_entry_for_row(row, schedule_map) is None and bool(
        row.get("hora_inicio") or row.get("hora_final")
    )


# ================================================================
#  LOGICA DE HORARIO Y TARDANZAS
# ================================================================

def _build_schedule_map(horarios):
    schedule_map = {}
    for row in horarios:
        dni      = str(row.get("dni_trabajador", "")).strip()
        day_name = str(row.get("dia_semana", "")).strip()
        entrada  = row.get("horario_entrada")
        if isinstance(entrada, str):
            try:
                entrada = datetime.strptime(entrada[:5], "%H:%M").time()
            except ValueError:
                entrada = None
        schedule_map.setdefault(dni, {})[day_name] = entrada
    return schedule_map


def _apply_late_flag(rows, schedule_map):
    for row in rows:
        parsed = _parse_date(row.get("fecha"))
        if not parsed:
            row["late"] = False
            continue
        dni       = str(row.get("dni", "")).strip()
        scheduled = schedule_map.get(dni, {}).get(DAY_NAME_MAP.get(parsed.weekday(), ""))
        actual    = _parse_time_value(row.get("hora_inicio"))
        if scheduled and actual:
            try:
                row["late"] = datetime.strptime(actual[:5], "%H:%M").time() > scheduled
            except ValueError:
                row["late"] = False
        else:
            row["late"] = False


# ================================================================
#  FILTROS
# ================================================================

def _filter_context(asistencias, trabajadores, selected_store_label, store_options, search_query):
    filtered_workers     = list(trabajadores)
    filtered_asistencias = list(asistencias)

    if selected_store_label != "Todas":
        store = store_options[selected_store_label]
        filtered_workers = [
            w for w in filtered_workers
            if str(w.get("id_sede", "")).strip() == str(store["id_tienda"]).strip()
        ]
        worker_ids = {w.get("dni", "") for w in filtered_workers}
        filtered_asistencias = [
            r for r in filtered_asistencias
            if str(r.get("dni", "")).strip() in worker_ids
        ]

    if search_query:
        q = search_query.lower()
        filtered_workers = [
            w for w in filtered_workers
            if q in f"{w.get('nombre_trabajador','')} {w.get('dni','')} {w.get('nombre_sede','')}".lower()
        ]
        worker_ids = {w.get("dni", "") for w in filtered_workers}
        filtered_asistencias = [
            r for r in filtered_asistencias
            if str(r.get("dni", "")).strip() in worker_ids
        ]

    return filtered_workers, filtered_asistencias


def _filter_context_with_worker(
    asistencias,
    trabajadores,
    selected_store_label,
    store_options,
    search_query,
    selected_worker_dni=None,
):
    filtered_workers, filtered_asistencias = _filter_context(
        asistencias,
        trabajadores,
        selected_store_label,
        store_options,
        search_query,
    )

    worker_dni = str(selected_worker_dni or "").strip()
    if worker_dni:
        filtered_workers = [
            w for w in filtered_workers
            if str(w.get("dni", "")).strip() == worker_dni
        ]
        filtered_asistencias = [
            r for r in filtered_asistencias
            if str(r.get("dni", "")).strip() == worker_dni
        ]

    return filtered_workers, filtered_asistencias


def _period_bounds(period_type, reference_date):
    reference_date = reference_date or date.today()
    if period_type == "Mes":
        start = date(reference_date.year, reference_date.month, 1)
        end   = date(reference_date.year, reference_date.month,
                     monthrange(reference_date.year, reference_date.month)[1])
        return start, end, _month_label(start)

    if reference_date.day <= 15:
        start = date(reference_date.year, reference_date.month, 1)
        end   = date(reference_date.year, reference_date.month, 15)
        label = f"{_month_label(start)} - Quincena 1"
    else:
        start = date(reference_date.year, reference_date.month, 16)
        end   = date(reference_date.year, reference_date.month,
                     monthrange(reference_date.year, reference_date.month)[1])
        label = f"{_month_label(start)} - Quincena 2"
    return start, end, label


def _collect_period_rows(asistencias, workers, start_date, end_date, schedule_map):
    workers_by_dni = {str(w.get("dni", "")).strip(): w for w in workers}
    rows = []
    for row in asistencias:
        parsed = _parse_date(row.get("fecha"))
        if not parsed or not (start_date <= parsed <= end_date):
            continue
        dni    = str(row.get("dni", "")).strip()
        worker = workers_by_dni.get(dni)
        if not worker:
            continue
        scheduled_entry = _scheduled_entry_for_row(row, schedule_map)
        if scheduled_entry is None and not _as_bool(row.get("justificado")):
            continue
        row_copy = dict(row)
        row_copy["nombre_trabajador"] = worker.get("nombre_trabajador", row_copy.get("nombre_trabajador", ""))
        row_copy["nombre_sede"]       = worker.get("nombre_sede",       row_copy.get("nombre_sede", ""))
        row_copy["cargo"]             = worker.get("area") or worker.get("cargo", "")
        rows.append(row_copy)

    _apply_late_flag(rows, schedule_map)
    return sorted(rows, key=lambda r: (r.get("fecha", ""), r.get("nombre_trabajador", "")))


def _build_attendance_excel(period_label, selected_store_label, search_query,
                            selected_worker_label, workers, rows, start_date, end_date):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from io import BytesIO
    except ImportError:
        return None

    def fmt_time(val):
        if not val or str(val).strip() in ("", "None", "--"):
            return ""
        text = str(val).strip()
        if "T" in text:
            text = text.split("T", 1)[-1]
        if " " in text:
            text = text.split(" ", 1)[-1]
        return text[:5]

    def parse_minutes(val):
        text = fmt_time(val)
        if not text:
            return None
        try:
            parsed = datetime.strptime(text[:5], "%H:%M").time()
            return parsed.hour * 60 + parsed.minute
        except ValueError:
            return None

    def calc_hours(row):
        entry_minutes = parse_minutes(row.get("hora_inicio"))
        exit_minutes = parse_minutes(row.get("hora_final"))
        if entry_minutes is None or exit_minutes is None:
            return "", ""

        exit_floor_minutes = (exit_minutes // 60) * 60
        total_minutes = exit_floor_minutes - entry_minutes

        has_break = bool(fmt_time(row.get("inicio_receso")) and fmt_time(row.get("final_receso")))
        if has_break:
            total_minutes -= 60

        total_minutes = max(total_minutes, 0)
        hours = total_minutes // 60
        minutes = total_minutes % 60
        entry_text = fmt_time(row.get("hora_inicio")) or "-"
        exit_text = fmt_time(row.get("hora_final")) or "-"
        exit_floor_text = f"{exit_floor_minutes // 60:02d}:00"
        calc_text = f"{exit_floor_text} - {entry_text}"
        if has_break:
            calc_text += " - 1:00 receso"
        calc_text += f" = {hours}:{minutes:02d}"
        return f"{hours}:{minutes:02d}", calc_text

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"

    title_fill = PatternFill("solid", fgColor="1e3a5f")
    title_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="dbeafe")
    header_font = Font(color="1e3a5f", bold=True)

    ws["A1"] = "Reporte de Asistencias"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    metadata = [
        ("Periodo", period_label),
        ("Rango", f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"),
        ("Tienda", selected_store_label if selected_store_label != "Todas" else "Todas las tiendas"),
        ("Persona", selected_worker_label if selected_worker_label != "Todas" else "Todas"),
        ("Busqueda", search_query or "-"),
        ("Generado", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    row_idx = 3
    for label, value in metadata:
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        row_idx += 1

    row_idx += 1
    headers = [
        "#",
        "Trabajador",
        "DNI",
        "Sede",
        "Fecha",
        "Entrada",
        "Ini. receso",
        "Fin receso",
        "Salida",
        "Horas",
        "Cálculo horas",
        "Tardanza",
        "Justificado",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    data_start = row_idx + 1
    for i, row in enumerate(rows, 1):
        horas, calculo = calc_hours(row)
        ws.append(
            [
                i,
                row.get("nombre_trabajador") or "-",
                row.get("dni") or "-",
                row.get("nombre_sede") or "-",
                row.get("fecha") or "-",
                fmt_time(row.get("hora_inicio")),
                fmt_time(row.get("inicio_receso")),
                fmt_time(row.get("final_receso")),
                fmt_time(row.get("hora_final")),
                horas,
                calculo,
                "Sí" if row.get("late") else "No",
                "Sí" if _as_bool(row.get("justificado")) else "No",
            ]
        )

    if ws.max_row >= data_start:
        ws.freeze_panes = f"A{data_start}"
        ws.auto_filter.ref = f"A{row_idx}:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ================================================================
#  PDF CON REPORTLAB
# ================================================================
def _build_attendance_pdf(period_label, selected_store_label, search_query,
                           workers, rows, start_date, end_date, tiendas=None,
                           schedule_map=None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from io import BytesIO
    from datetime import timedelta

    # â”€â”€ Paleta â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    C_BLACK      = colors.HexColor("#0f172a")
    C_DARK       = colors.HexColor("#1e293b")
    C_MID        = colors.HexColor("#64748b")
    C_LIGHT      = colors.HexColor("#f1f5f9")
    C_WHITE      = colors.white
    C_HEADER_BG  = colors.HexColor("#1e3a5f")   # azul oscuro â€” header tabla
    C_WEEK_BG    = colors.HexColor("#dbeafe")    # azul claro â€” cabecera semana
    C_WEEK_TEXT  = colors.HexColor("#1e40af")    # azul semana texto
    C_ROW_ALT    = colors.HexColor("#f8fafc")    # fila alternada
    C_ACCENT     = colors.HexColor("#2563eb")
    C_GREEN_TEXT = colors.HexColor("#166534")
    C_JUST_BLUE  = colors.HexColor("#5b8def")
    C_JUST_BLUE_TEXT = colors.HexColor("#1d4ed8")
    C_RED_TEXT   = colors.HexColor("#991b1b")
    C_RED_BG     = colors.HexColor("#dc2626")
    C_SAT_BG     = colors.HexColor("#fef3c7")    # amarillo â€” tardanza
    C_ABSENT_BG  = colors.HexColor("#f1f5f9")    # gris â€” falta

    # â”€â”€ Estilos â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    def sty(name, font="Helvetica", size=8, color=None, bold=False,
            align=None, leading=10, italic=False, back_color=None):
        fn = "Helvetica-BoldOblique" if bold and italic else \
             "Helvetica-Bold" if bold else \
             "Helvetica-Oblique" if italic else font
        kw = dict(fontName=fn, fontSize=size,
                  textColor=color or C_BLACK, leading=leading)
        if align:
            kw["alignment"] = align
        if back_color is not None:
            kw["backColor"] = back_color
        return ParagraphStyle(name, **kw)

    sty_title    = sty("title",  bold=True,  size=16, leading=20)
    sty_subtitle = sty("sub",    size=8,     color=C_DARK, leading=11)
    sty_footer   = sty("foot",   size=6.5,   color=C_MID, align=TA_RIGHT, leading=9)
    sty_ml       = sty("ml",     bold=True,  size=7.5)
    sty_mv       = sty("mv",     size=7.5,   color=C_DARK)
    sty_th       = sty("th",     bold=True,  size=7.5, color=C_WHITE, align=TA_CENTER)
    sty_th_l     = sty("thl",    bold=True,  size=7.5, color=C_WHITE)
    sty_week_hd  = sty("wk",     bold=True,  size=8,   color=C_WEEK_TEXT)
    sty_cell     = sty("cell",   size=7,     color=C_BLACK)
    sty_cell_c   = sty("cellc",  size=7,     color=C_BLACK, align=TA_CENTER)
    sty_bold     = sty("bold",   bold=True,  size=7,   color=C_BLACK)
    sty_late     = sty("late",   bold=True,  size=6.5, color="#92400e", align=TA_CENTER, back_color=C_SAT_BG)
    sty_ok       = sty("ok",     bold=True,  size=6.5, color=C_GREEN_TEXT, align=TA_CENTER, back_color=colors.HexColor("#dcfce7"))
    sty_justified = sty("just",   bold=True,  size=6.5, color=C_JUST_BLUE_TEXT, align=TA_CENTER, back_color=colors.HexColor("#dbeafe"))
    sty_absent   = sty("abs",    bold=True,  size=6.5, color=C_WHITE, align=TA_CENTER, back_color=C_RED_BG)
    sty_missing  = sty("miss",   bold=True,  size=6.5, color=C_MID, align=TA_CENTER, back_color=C_ABSENT_BG)
    sty_num      = sty("num",    size=7,     color=C_MID, align=TA_CENTER)

    # â”€â”€ Helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    def fmt_time(val):
        if not val or str(val).strip() in ("", "None", "--"):
            return ""
        s = str(val).strip()
        if "T" in s: s = s.split("T")[-1]
        if " " in s: s = s.split(" ")[-1]
        return s[:5]

    def cell_marcaciones(row_data):
        """Devuelve un Paragraph con hasta 4 marcaciones apiladas."""
        if row_data is None:
            return Paragraph("FALTA", sty_absent)
        e  = fmt_time(row_data.get("hora_inicio"))
        r1 = fmt_time(row_data.get("inicio_receso"))
        r2 = fmt_time(row_data.get("final_receso"))
        s  = fmt_time(row_data.get("hora_final"))
        is_late = bool(row_data.get("late"))
        is_justified = _as_bool(row_data.get("justificado"))
        scheduled_entry = _scheduled_entry_for_row(row_data, schedule_map or {})

        parts = []
        if is_justified and scheduled_entry is None:
            parts.append("JUSTIFICADO")
        if e:
            parts.append(e)
        if r1:
            parts.append(r1)
        if r2:
            parts.append(r2)
        if s:
            parts.append(s)
        if not parts:
            return Paragraph("â€”", sty_missing)
        if is_justified and scheduled_entry is None:
            return Paragraph("<br/>".join(parts), sty_justified)
        return Paragraph("<br/>".join(parts), sty_late if is_late else sty_ok)

    # â”€â”€ Organizar datos â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Ã­ndice: dni â†’ fecha â†’ row
    from collections import defaultdict
    att_index = defaultdict(dict)
    for row in rows:
        try:
            d = date.fromisoformat(str(row.get("fecha", ""))[:10])
        except ValueError:
            continue
        dni = str(row.get("dni", "")).strip()
        att_index[dni][d] = row

    # semanas dentro del periodo (lunes a sÃ¡bado)
    def get_weeks(start, end):
        weeks = []
        cursor = start - timedelta(days=start.weekday())  # lunes de la semana
        week_num = 1
        while cursor <= end:
            days = []
            for i in range(6):  # lunes=0 ... sÃ¡bado=5
                d = cursor + timedelta(days=i)
                if start <= d <= end:
                    days.append(d)
            if days:
                weeks.append((week_num, days))
                week_num += 1
            cursor += timedelta(days=7)
        return weeks

    weeks = get_weeks(start_date, end_date)

    # â”€â”€ Stats generales â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    generated_at  = datetime.now().strftime("%d/%m/%Y %H:%M")
    store_label   = selected_store_label if selected_store_label != "Todas" else "Todas las tiendas"
    tardanzas     = sum(1 for r in rows if r.get("late"))
    puntual_count = len(rows) - tardanzas

    # â”€â”€ Documento â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=12*mm,  bottomMargin=12*mm,
    )
    story = []

    # â”€â”€ TÃTULO â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    story.append(Paragraph("Reporte de Asistencias", sty_title))
    story.append(Paragraph(
        f"Periodo: <b>{period_label}</b>  Â·  "
        f"Rango: {start_date.strftime('%d/%m/%Y')} â€” {end_date.strftime('%d/%m/%Y')}  Â·  "
        f"Generado: {generated_at}", sty_subtitle))
    story.append(HRFlowable(width="100%", thickness=1.2, color=C_ACCENT, spaceAfter=2*mm))

    # â”€â”€ METADATOS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    meta = [
        [Paragraph("<b>Tienda</b>",       sty_ml), Paragraph(store_label,        sty_mv),
         Paragraph("<b>Trabajadores</b>", sty_ml), Paragraph(str(len(workers)),  sty_mv),
         Paragraph("<b>Registros</b>",    sty_ml), Paragraph(str(len(rows)),     sty_mv)],
        [Paragraph("<b>Busqueda</b>",     sty_ml), Paragraph(search_query or "â€”", sty_mv),
         Paragraph("<b>A tiempo</b>",     sty_ml), Paragraph(str(puntual_count), sty_mv),
         Paragraph("<b>Tardanzas</b>",    sty_ml), Paragraph(str(tardanzas),     sty_mv)],
    ]
    meta_t = Table(meta, colWidths=[22*mm, 68*mm, 28*mm, 22*mm, 24*mm, 22*mm])
    meta_t.setStyle(TableStyle([
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [C_WHITE, C_LIGHT]),
        ("BOX",            (0,0), (-1,-1), 0.5, C_MID),
        ("INNERGRID",      (0,0), (-1,-1), 0.25, C_MID),
        ("TOPPADDING",     (0,0), (-1,-1), 2),
        ("BOTTOMPADDING",  (0,0), (-1,-1), 2),
        ("LEFTPADDING",    (0,0), (-1,-1), 4),
        ("RIGHTPADDING",   (0,0), (-1,-1), 4),
        ("VALIGN",         (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(meta_t)
    story.append(Spacer(1, 3*mm))

    # â”€â”€ LEYENDA â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    leyenda_data = [[
        Paragraph("<b>Leyenda de colores:</b>", sty_ml),
        Paragraph("Puntual", sty_ok),
        Paragraph("Tardanza", sty_late),
        Paragraph("Justificado", sty_justified),
        Paragraph("Falta", sty_absent),
        Paragraph("Sin marca", sty_missing),
    ]]
    ley_t = Table(leyenda_data, colWidths=[34*mm, 32*mm, 34*mm, 34*mm, 28*mm, 30*mm])
    ley_t.setStyle(TableStyle([
        ("BOX",           (0,0), (-1,-1), 0.4, C_MID),
        ("TOPPADDING",    (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
        ("BACKGROUND",    (0,0), (-1,-1), C_LIGHT),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(ley_t)
    story.append(Spacer(1, 4*mm))

    # â”€â”€ TABLAS POR SEMANA â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    DAY_NAMES = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado"]

    for week_num, week_days in weeks:
        # encabezado de semana
        week_start_str = week_days[0].strftime("%d/%m")
        week_end_str   = week_days[-1].strftime("%d/%m")
        story.append(Paragraph(
            f"Semana {week_num}  Â·  {week_start_str} â€” {week_end_str}",
            sty_week_hd,
        ))

        # fechas del header (solo los dÃ­as presentes en la semana)
        day_headers = [
            f"{DAY_NAMES[d.weekday()]}\n{d.strftime('%d/%m')}"
            for d in week_days
        ]

        # anchos: NÂ° + Trabajador + dÃ­as
        n_days    = len(week_days)
        day_w     = (246 - 8 - 42) / n_days  # distribuir espacio disponible
        col_ws    = [8*mm, 42*mm] + [day_w*mm] * n_days

        # header row
        header_row = [
            Paragraph("NÂ°",         sty_th),
            Paragraph("Trabajador", sty_th_l),
        ] + [Paragraph(h.replace("\n", "<br/>"), sty_th) for h in day_headers]

        # filas de trabajadores
        table_rows = [header_row]
        style_cmds = [
            ("BACKGROUND",    (0,0), (-1,0),  C_HEADER_BG),
            ("BOX",           (0,0), (-1,-1), 0.5, C_MID),
            ("INNERGRID",     (0,0), (-1,-1), 0.25, C_MID),
            ("TOPPADDING",    (0,0), (-1,-1), 2),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("LEFTPADDING",   (0,0), (-1,-1), 3),
            ("RIGHTPADDING",  (0,0), (-1,-1), 3),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_WHITE, C_ROW_ALT]),
        ]

        for row_idx, worker in enumerate(workers, 1):
            dni    = str(worker.get("dni", "")).strip()
            nombre = _safe_text(worker.get("nombre_trabajador", ""))
            worker_att = att_index.get(dni, {})

            cells = [
                Paragraph(str(row_idx), sty_num),
                Paragraph(nombre,       sty_bold),
            ]

            for col_idx, day in enumerate(week_days, 2):
                row_data = worker_att.get(day)
                cells.append(cell_marcaciones(row_data))

            table_rows.append(cells)

        week_table = Table(table_rows, colWidths=col_ws, repeatRows=1)
        week_table.setStyle(TableStyle(style_cmds))

        story.append(week_table)
        story.append(Spacer(1, 5*mm))

    # â”€â”€ PIE â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    story.append(HRFlowable(width="100%", thickness=0.4, color=C_MID))
    story.append(Paragraph(
        f"Generado el {generated_at}  |  {len(workers)} trabajadores  |  "
        f"{len(rows)} registros  |  {puntual_count} a tiempo  |  {tardanzas} tardanzas",
        sty_footer,
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()

# ================================================================
#  HTML TABLA SEMANAL
# ================================================================

def _cell_html(attendance_row, scheduled_entry=None, current_date=None, cell_date=None):
    if scheduled_entry is None:
        if attendance_row:
            entrada = attendance_row.get("hora_inicio") or "-"
            salida  = attendance_row.get("hora_final")  or "-"
            if _as_bool(attendance_row.get("justificado")):
                return (
                    '<div style="display:flex;flex-direction:column;gap:.12rem;line-height:1.15;'
                    'background:#dbeafe;padding:.4rem .45rem;border-radius:10px;">'
                    '<div style="font-family:Space Mono,monospace;font-size:.66rem;'
                    'color:#1d4ed8;font-weight:700;text-transform:uppercase;letter-spacing:.05em;">'
                    'Justificado'
                    '</div>'
                    f'<span style="font-family:Space Mono,monospace;font-size:.78rem;color:#1d4ed8;font-weight:700;">{entrada}</span>'
                    f'<span style="font-family:Space Mono,monospace;font-size:.78rem;color:#1d4ed8;font-weight:700;">{salida}</span>'
                    "</div>"
                )
            return (
                '<div style="display:flex;flex-direction:column;gap:.12rem;line-height:1.15;'
                'background:#dbeafe;padding:.4rem .45rem;border-radius:10px;">'
                '<div style="font-family:Space Mono,monospace;font-size:.66rem;'
                'color:#1d4ed8;font-weight:700;text-transform:uppercase;letter-spacing:.05em;">'
                'Justificable'
                '</div>'
                f'<span style="font-family:Space Mono,monospace;font-size:.78rem;color:#1d4ed8;font-weight:700;">{entrada}</span>'
                f'<span style="font-family:Space Mono,monospace;font-size:.78rem;color:#1d4ed8;font-weight:700;">{salida}</span>'
                "</div>"
            )
        return (
            '<div style="display:flex;align-items:center;justify-content:center;'
            'min-height:2.3rem;padding:.35rem .45rem;border-radius:10px;'
            'background:#e5e7eb;color:#475569;font-family:Space Mono,monospace;'
            'font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.04em;">'
            'No viene'
            '</div>'
        )

    if not attendance_row:
        if current_date and cell_date and cell_date > current_date:
            return '<div style="min-height:2.3rem;"></div>'
        return (
            '<div style="display:flex;align-items:center;justify-content:center;'
            'min-height:2.3rem;padding:.35rem .45rem;border-radius:10px;'
            'background:#fff1f2;color:#991b1b;font-family:Space Mono,monospace;'
            'font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.04em;">'
            'Falta'
            '</div>'
        )

    entrada = attendance_row.get("hora_inicio") or "-"
    salida  = attendance_row.get("hora_final")  or "-"
    on_time = _is_on_time_or_early(attendance_row.get("hora_inicio"), scheduled_entry)
    color = "#64748b" if on_time is None else ("#166534" if on_time else "#92400e")
    bg = "#dcfce7" if on_time else "#f59e0b"
    return (
        f'<div style="display:flex;flex-direction:column;gap:.12rem;line-height:1.15;'
        f'background:{bg};padding:.4rem .45rem;border-radius:10px;">'
        f'<span style="font-family:Space Mono,monospace;font-size:.78rem;color:{color};font-weight:700;">{entrada}</span>'
        f'<span style="font-family:Space Mono,monospace;font-size:.78rem;color:#166534;font-weight:700;">{salida}</span>'
        "</div>"
    )


def _render_weekly_matrix(workers, attendance_map, days, schedule_map):
    rows_html = []
    spanish_days = {
        0: "Lunes",
        1: "Martes",
        2: "Miércoles",
        3: "Jueves",
        4: "Viernes",
        5: "Sábado",
    }
    today = date.today()
    for worker in workers:
        dni   = worker.get("dni", "")
        w_att = attendance_map.get(dni, {})
        cells = "".join(
            f"<td>{_cell_html(w_att.get(d), schedule_map.get(dni, {}).get(DAY_NAME_MAP.get(d.weekday(), '')), today, d)}</td>"
            for d in days
        )
        rows_html.append(f"""
        <tr>
          <td class="worker-td">
            <div class="worker-name">{worker.get('nombre_trabajador','-')}</div>
            <div class="worker-meta">DNI {dni} Â· {worker.get('nombre_sede','-')}</div>
          </td>
        {cells}
        </tr>""")

    header_cells = "".join(
        f"<th>{spanish_days.get(d.weekday(), d.strftime('%A'))}<br>"
        f"<span style='font-weight:400;font-size:.6rem;'>{d.strftime('%d/%m')}</span></th>"
        for d in days
    )
    return f"""
    <html><head><style>
      body{{margin:0;font-family:'DM Sans',sans-serif;}}
      .wrap{{overflow-x:auto;border:1px solid #e2e8f0;border-radius:12px;
             background:#fff;box-shadow:0 1px 6px rgba(15,23,42,.04);}}
      .at-table{{width:100%;border-collapse:collapse;font-size:.85rem;}}
      .at-table thead th{{text-align:left;padding:.85rem .9rem;background:#f8fafc;
        color:#64748b;font-family:'Space Mono',monospace;font-size:.68rem;
        text-transform:uppercase;letter-spacing:.1em;border-bottom:1px solid #e2e8f0;
        white-space:nowrap;}}
      .at-table tbody tr{{border-bottom:1px solid #f1f5f9;}}
      .at-table tbody tr:hover{{background:#fafbff;}}
      .at-table td{{padding:1rem .9rem;vertical-align:middle;color:#1e293b;}}
      .at-table td.worker-td{{min-width:180px;}}
      .worker-name{{font-weight:700;color:#0f172a;font-size:.85rem;}}
      .worker-meta{{font-size:.69rem;color:#64748b;margin-top:.15rem;}}
      .m-empty{{color:#94a3b8;font-family:'Space Mono',monospace;font-size:.78rem;}}
      .m-no-come{{display:flex;align-items:center;justify-content:center;min-height:2.3rem;
        padding:.35rem .45rem;border-radius:10px;background:#e5e7eb;color:#475569;
        font-family:'Space Mono',monospace;font-size:.72rem;font-weight:700;
        text-transform:uppercase;letter-spacing:.04em;}}
    </style></head>
    <body><div class="wrap">
      <table class="at-table">
        <thead><tr><th>Trabajador</th>{header_cells}</tr></thead>
        <tbody>{''.join(rows_html)}</tbody>
      </table>
    </div></body></html>"""


def _render_pending_justifications(api, rows, schedule_map):
    pending = [row for row in rows if _is_justificable_row(row, schedule_map)]
    if not pending:
        st.caption("No hay justificaciones pendientes.")
        return

    st.markdown(
        """
        <div style="margin-top:0.75rem;margin-bottom:0.35rem;
                    font-family:'Space Mono',monospace;font-size:0.72rem;
                    text-transform:uppercase;letter-spacing:0.08em;color:#64748b;">
            Justificaciones pendientes
        </div>
        """,
        unsafe_allow_html=True,
    )

    for row in pending:
        left, right = st.columns([5, 1])
        left.caption(
            f"{row.get('nombre_trabajador', '-')} · DNI {row.get('dni', '-')} · {row.get('fecha', '-')}"
        )
        if right.button(
            "✓",
            key=f"justify_{row.get('doc_id') or row.get('ruta') or row.get('dni')}_{row.get('fecha', '')}",
            help="Marcar como justificado",
            use_container_width=True,
        ):
            try:
                api.update_document(
                    api.ATTENDANCE_COLLECTION,
                    row.get("doc_id") or row.get("ruta"),
                    {"justificado": True},
                    key_field="id_asistencia",
                )
                api.invalidate_collection_cache(api.ATTENDANCE_COLLECTION)
                st.rerun()
            except Exception as exc:
                st.error(f"No se pudo justificar el registro: {exc}")


# ================================================================
#  VISTA PRINCIPAL
# ================================================================

def render_resumen(api=None):
    if api is None:
        st.error("Falta el contexto de la app.")
        return

    st.markdown("""
    <div style="display:flex;align-items:center;gap:0.75rem;margin-bottom:0.25rem;">
        <div style="font-size:1.4rem;color:#2563eb;">¬</div>
        <div>
            <div style="font-family:'Space Mono',monospace;font-size:1.25rem;
                        letter-spacing:0.03em;color:#000000;line-height:1.1;">
                Asistencias
            </div>
            <div style="font-size:0.78rem;color:#000000;font-family:'DM Sans',sans-serif;
                        margin-top:0.15rem;">
                Vista semanal por trabajador.
            </div>
        </div>
    </div>
    <hr style="margin:0.75rem 0 1.5rem;border-color:#dde1ea;">
    """, unsafe_allow_html=True)

    asistencias, trabajadores, tiendas, horarios = api.get_resumen_dashboard()

    store_options = {"Todas": None}
    store_options.update({
        f"{s['nombre_tienda']} Â· {s['id_tienda']}": s for s in tiendas
    })

    if "resumen_week" not in st.session_state:
        all_dates = [_parse_date(r.get("fecha")) for r in asistencias if _parse_date(r.get("fecha"))]
        st.session_state["resumen_week"] = max(all_dates) if all_dates else date.today()

    current_start = _week_start(st.session_state["resumen_week"])
    week_end = current_start + timedelta(days=6)
    week_days = [current_start + timedelta(days=i) for i in range(6)]

    # Filtros y navegación semanal
    sc1, sc2, sc3, sc4, sc5 = st.columns([2.2, 1.3, 1.8, 0.75, 0.75])
    search_query = sc1.text_input(
        "Buscar", placeholder="Nombre, DNI o sede...",
        key="resumen_search", label_visibility="collapsed",
    )
    selected_store_label = sc2.selectbox(
        "Tienda", options=list(store_options.keys()),
        index=0, key="resumen_store_filter", label_visibility="collapsed",
    )
    worker_options = ["Todas"] + [
        f"{w.get('nombre_trabajador', w.get('nombre', '-'))} · {w.get('dni', '-')}"
        for w in trabajadores
    ]
    worker_labels = {opt: (None if opt == "Todas" else opt.split(" · ", 1)[-1]) for opt in worker_options}
    selected_worker_label = sc3.selectbox(
        "Persona", options=worker_options,
        index=0, key="resumen_worker_filter", label_visibility="collapsed",
    )
    if sc4.button("⬅", use_container_width=True, key="resumen_prev_button", help="Semana anterior"):
        st.session_state["resumen_week"] = current_start - timedelta(days=7)
        st.rerun()
    if sc5.button("➡", use_container_width=True, key="resumen_next_button", help="Semana siguiente"):
        st.session_state["resumen_week"] = current_start + timedelta(days=7)
        st.rerun()

    schedule_map = _build_schedule_map(horarios)
    _apply_late_flag(asistencias, schedule_map)

    filtered_workers, filtered_asistencias = _filter_context_with_worker(
        asistencias,
        trabajadores,
        selected_store_label,
        store_options,
        search_query,
        worker_labels.get(selected_worker_label),
    )

    # Tarjeta de rango de fechas
    range_label = (
        f"{current_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')}"
    )
    st.markdown(
        f"""
        <div style="
            border: 1px solid #dbe4ee;
            border-radius: 14px;
            background: #ffffff;
            box-shadow: 0 4px 18px rgba(15, 23, 42, 0.06);
            padding: 1rem 1.1rem;
            margin: 0.2rem 0 0.75rem;
        ">
            <div style="
                font-family: 'Space Mono', monospace;
                font-size: 0.65rem;
                letter-spacing: 0.12em;
                text-transform: uppercase;
                color: #64748b;
                margin-bottom: 0.25rem;
            ">Rango de fechas</div>
            <div style="
                font-family: 'Space Mono', monospace;
                font-size: 1rem;
                color: #0f172a;
                white-space: nowrap;
                overflow-x: auto;
                padding-bottom: 0.1rem;
            ">{range_label}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)
    st.caption(f"Semana: `{_week_label(current_start)}`")

    st.markdown(
        """
        <div style="display:flex;flex-wrap:wrap;gap:0.5rem;margin:0.3rem 0 0.9rem;">
            <span style="background:#dcfce7;color:#166534;border-radius:999px;padding:0.28rem 0.6rem;
                         font-family:'Space Mono',monospace;font-size:0.68rem;font-weight:700;">
                Puntual
            </span>
            <span style="background:#fef3c7;color:#92400e;border-radius:999px;padding:0.28rem 0.6rem;
                         font-family:'Space Mono',monospace;font-size:0.68rem;font-weight:700;">
                Tardanza
            </span>
            <span style="background:#dbeafe;color:#1d4ed8;border-radius:999px;padding:0.28rem 0.6rem;
                         font-family:'Space Mono',monospace;font-size:0.68rem;font-weight:700;">
                Justificado
            </span>
            <span style="background:#fff1f2;color:#991b1b;border-radius:999px;padding:0.28rem 0.6rem;
                         font-family:'Space Mono',monospace;font-size:0.68rem;font-weight:700;">
                Falta
            </span>
            <span style="background:#e5e7eb;color:#475569;border-radius:999px;padding:0.28rem 0.6rem;
                         font-family:'Space Mono',monospace;font-size:0.68rem;font-weight:700;">
                Sin marca
            </span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    attendance_map = {}
    for row in filtered_asistencias:
        parsed = _parse_date(row.get("fecha"))
        if not parsed or not (current_start <= parsed <= week_end):
            continue
        attendance_map.setdefault(str(row.get("dni","")).strip(), {})[parsed] = row

    if not filtered_workers:
        st.info("No hay trabajadores o asistencias para este rango.")
        return

    # Exportar PDF
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    ex1, ex2, ex3, ex4 = st.columns([1.1, 1.1, 1.0, 1.0])

    export_period = ex1.selectbox(
        "Periodo PDF", ["Mes", "Quincena"],
        index=0, key="resumen_export_period",
    )
    export_reference = ex2.date_input(
        "Referencia PDF",
        value=st.session_state.get("resumen_week", date.today()),
        key="resumen_export_reference",
    )
    st.caption("El PDF y el Excel respetan los filtros activos de tienda, busqueda y persona.")

    export_start, export_end, period_label = _period_bounds(export_period, export_reference)
    export_rows = _collect_period_rows(
        filtered_asistencias, filtered_workers,
        export_start, export_end, schedule_map,
    )
    pending_justifications = [
        row for row in filtered_asistencias
        if _is_justificable_row(row, schedule_map)
    ]

    if export_rows:
        pdf_bytes = _build_attendance_pdf(
            period_label         = period_label,
            selected_store_label = selected_store_label,
            search_query         = search_query,
            workers              = filtered_workers,
            rows                 = export_rows,
            start_date           = export_start,
            end_date             = export_end,
            tiendas              = tiendas,
            schedule_map         = schedule_map,
        )
        pdf_name = (
            f"asistencias_{_sanitize_filename(period_label)}_"
            f"{_sanitize_filename(selected_store_label)}.pdf"
        )
        ex3.download_button(
            "Exportar PDF",
            data=pdf_bytes,
            file_name=pdf_name,
            mime="application/pdf",
            use_container_width=True,
        )
        excel_bytes = _build_attendance_excel(
            period_label         = period_label,
            selected_store_label = selected_store_label,
            search_query         = search_query,
            selected_worker_label = selected_worker_label,
            workers              = filtered_workers,
            rows                 = export_rows,
            start_date           = export_start,
            end_date             = export_end,
        )
        if excel_bytes:
            excel_name = (
                f"asistencias_{_sanitize_filename(period_label)}_"
                f"{_sanitize_filename(selected_store_label)}.xlsx"
            )
            ex4.download_button(
                "Exportar Excel",
                data=excel_bytes,
                file_name=excel_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            ex4.info("Instala openpyxl para exportar Excel.")
        if pending_justifications:
            ex3.caption(
                f"{len(pending_justifications)} registro(s) pendientes de justificar no se exportan aún."
            )
    else:
        if pending_justifications:
            ex3.warning("Hay registros pendientes de justificar; todavía no se exportan.")
        else:
            ex3.info("Sin registros para el periodo seleccionado.")

    with st.expander("Justificaciones pendientes", expanded=False):
        _render_pending_justifications(api, filtered_asistencias, schedule_map)

    # Tabla semanal
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    table_html = _render_weekly_matrix(filtered_workers, attendance_map, week_days, schedule_map)
    components.html(table_html,
                    height=max(320, 92 + len(filtered_workers) * 88),
                    scrolling=True)

